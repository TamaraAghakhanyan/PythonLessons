import openpyxl

#to impport workbook on python:
workbook = openpyxl.load_workbook("/Users/tamaraaghakhanyan/Downloads/Employees.xlsx")

#to get sheet names:
print(workbook.sheetnames)

# to get the active sheet in the workbook:
print(workbook.active)

# to reference a sheet by its name:
sheet = workbook["EmployeeData"]
print(sheet)

# to create a new sheet in the workbook:
print(workbook.create_sheet("TestS"))

# to save the new changes:
print(workbook.save("/Users/tamaraaghakhanyan/Downloads/Employees.xlsx"))

# to remove a sheet from the workbook:
sheet = workbook["TestSheet"]
print(workbook.remove(sheet))

# OR

del workbook["TestSheet"]

print(workbook.save("/Users/tamaraaghakhanyan/Downloads/Employees.xlsx"))

#Closing a workbook
workbook.close()

#Getting the title of a sheet
print(sheet.title)

# Renaming the default sheet to 'Employees'
sheet = workbook['Sheet']

sheet.title = 'Employees'

# to get the active cell in the sheet:
print(sheet.active_cell)

#to get the dimensions of the sheet (the array populated with data):
print(sheet.dimensions)

# basic parameters of the sheeet:
print(sheet.sheet_format)

print(sheet.sheet_properties)

# returns the state of the sheet(visible or hidden inside the workbook):
print(sheet.sheet_state)


print(sheet.sheet_view)

# getting the numbers of the rows/columns inside the sheet:
print(sheet.max_row)

print(sheet.max_column)

# return the rows inside the sheet as tuples:
for i in sheet.values:
    print(i)

# getting the value of the cell using sheet coordinates:
sheet = workbook["EmployeeData"]
print(sheet["B7"].value)

# OR

print(sheet.cell(row = 6, column = 2).value)

# referencing a particular cell in the sheet:
cell = sheet["B9"]
print(cell)

print(cell.row)

print(cell.column_letter)

print(cell.coordinates)

#getting the data type of the value inside the cell:
print(cell.data_type)

'''
TYPE_STRING = 's'
    TYPE_FORMULA = 'f'
    TYPE_NUMERIC = 'n'
    TYPE_BOOL = 'b'
    TYPE_NULL = 'n'
    TYPE_INLINE = 'inlineStr'
    TYPE_ERROR = 'e'
    TYPE_FORMULA_CACHE_STRING = 'str'
'''

# checking out encoding format for the cell:
print(cell.encoding)

# setting a certain value for a cell:
cell = sheet["B2"]
cell.value = "David"
workbook.save("/Users/tamaraaghakhanyan/Downloads/Employees.xlsx")

# getting the parent worksheet of the cell:
print(cell.parent)

# Working with cell styles
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Color


#open workbook:
workbook = openpyxl.load_workbook("/Users/tamaraaghakhanyan/Downloads/Employees.xlsx")

# accessing the sheet:
sheet = workbook["EmployeeData"]

# accessing the cell:
cell = sheet["B8"]

# Create a Color object with the desired color (2 represents RED color index)
font_color = Color(rgb="FF0000")

# Create a Font object with the desired properties, including the color
font = Font(color = 2, bold = True, italic = True)


# Apply the Font to a cell (e.g., B8)
cell = sheet['B8']
cell.font = font

# Save the workbook
workbook.save("/Users/tamaraaghakhanyan/Downloads/Employees.xlsx")

# Adding background color to the cell:
from openpyxl.styles import PatternFill
fill = PatternFill(fill_type = "solid", bgColor = "00FFFF99")

# Apply the Pattern to a cell
cell.fill = fill

# Save the workbook
workbook.save("/Users/tamaraaghakhanyan/Downloads/Employees.xlsx")

# borders:
from openpyxl.styles import Border, Side

border = Border(left = Side(border_style = 'double', color = '322FEC'), right = Side(border_style = 'double', color = '322FEC'), top = Side(border_style = 'double', color = '322FEC'), bottom = Side(border_style = 'double', color = '322FEC'))
cell.border = border
workbook.save("/Users/tamaraaghakhanyan/Downloads/Employees.xlsx")

from openpyxl.styles import Alignment
align = Alignment(horizontal = "left")
cell.alignment = align